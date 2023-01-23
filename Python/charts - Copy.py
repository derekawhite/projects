#python -m pip install spacepy
# python -m pip install --trusted-host pypi.python.org --trusted-host pypi.org --trusted-host files.pythonhosted.org xlsxwriter
# python -m pip install --trusted-host pypi.python.org --trusted-host pypi.org --trusted-host files.pythonhosted.org pandas
# python -m pip install --trusted-host pypi.python.org --trusted-host pypi.org --trusted-host files.pythonhosted.org xlrd
# python -m pip install --trusted-host pypi.python.org --trusted-host pypi.org --trusted-host files.pythonhosted.org openpyxl

from html.parser import HTMLParser
from html.entities import name2codepoint
import urllib.request
from datetime import date, timedelta
import datetime
import functools
import sys
import os
from xlsxwriter.workbook import Workbook 
import openpyxl

def remove_non_ascii(string):
    return ''.join(char for char in string if ord(char) < 128)

class song:
    def __init__(self, title, artist, position, peak, weeks, chartdate, enddate):
        self.title = remove_non_ascii(title)
        self.artist = remove_non_ascii(artist)
        self.position = position
        self.peak = peak        
        self.weeks = weeks
        self.startdate = chartdate
        self.enddate = enddate

class artist:
    def __init__(self, artist):
        self.artist = remove_non_ascii(artist)       
        self.weeks = 1

class MyHTMLParser(HTMLParser):
    songs = []
    inArtist = False
    inTitle = False
    inPosition = False
    inPeak = False    
    inWks = False

    def clear(self):
        self.songs.clear();

    def handle_comment(self, data):
        self.inPeak = False    
        self.inWks = False
        comment = data.strip()
        if comment == 'Peak Position':
            self.inPeak = True
        if comment == 'Wks':
            self.inWks = True

    def handle_starttag(self, tag, attrs):
        if len(attrs) >= 1 and len(attrs[0]) > 1:
            if attrs[0][0] == 'class':
                if attrs[0][1] == 'title':
                    self.inTitle = True
                if attrs[0][1] == 'artist':
                    self.inArtist = True
                if attrs[0][1] == 'position':
                    self.inPosition = True

    def handle_data(self, data):       
        string = data.strip()
        if len(string) > 0:
            if self.inTitle:
                self.inTitle = False
                self.Title = string
            if self.inPosition:
                self.Position = int(string)
                self.inPosition = False
            if self.inArtist:
                self.inArtist = False
                self.Artist = string
            if self.inPeak:
                self.Peak = int(string)
                self.inPeak = False
            if self.inWks:
                self.Weeks = int(string)
                self.inWks = False
                self.appendsong()

    def findsong(self, songtofind):
        for index, s in enumerate(self.songs):
            if s.title == songtofind.title and s.artist == songtofind.artist:
                return s
        return None

    def appendsong(self):
        newsong = song(self.Title, self.Artist, self.Position, self.Peak, self.Weeks, self.chartdate, self.chartdate)

        if bAll:
            self.songs.append(newsong) 
        else:            
            oldsong = self.findsong(newsong)
            if (oldsong != None):
                if oldsong.position < newsong.position:
                    newsong.position = oldsong.position 
                newsong.startdate = oldsong.startdate
                newsong.enddate = self.chartdate
                self.songs.remove(oldsong) 
                self.songs.append(newsong) 
            else:
                self.songs.append(newsong) 

def comparesongs(item1, item2):
    if item1.artist == item2.artist:
        if item1.title == item2.title:
            return 1 if item1.position > item2.position else -1
        else:
            return 1 if item1.title > item2.title else -1
    else:
        return 1 if item1.artist > item2.artist else -1    

def writeartists(artists):
    env = os.getenv("USERPROFILE")
    dirroot = f"{env}\\Dropbox (Personal)"
    if not os.path.isdir(dirroot):
        dirroot = f"{env}\\Dropbox"
        dirname = f"{dirroot}\\Charts\\Artists"
        filename = f"{dirname}\\Artists.xlsx"

    os.makedirs(dirname, exist_ok=True)

    while os.path.isfile(filename):       
        try:
            os.remove(filename)
            removed = True
        except:
            input(f"File {filename} not accessible. Close file and press enter to retry: ")
        
    workbook = Workbook(filename)
    worksheet = workbook.add_worksheet()
    format1 = workbook.add_format({'num_format': 'dd/mm/yyyy'})
    format2 = workbook.add_format()
    format2.set_bold()

    worksheet.freeze_panes(1, 0)

    col = 0
    worksheet.write (0, col, "WEEKS", format2);
    worksheet.set_column(col, col, 8)
    worksheet.write (0, col:=col+1, "ARTIST", format2);
    worksheet.set_column(col, col, 50)

    for i, a in enumerate(artists):
        col = 0
        worksheet.write(i+1, col, a.weeks)
        worksheet.write(i+1, col:=col+1, a.artist)

    workbook.close()

    print (f"output file:\n{filename}") 


def writefile():
    parser.songs.sort(key=functools.cmp_to_key(comparesongs))
    env = os.getenv("USERPROFILE")
    dirroot = f"{env}\\Dropbox (Personal)"
    if not os.path.isdir(dirroot):
        dirroot = f"{env}\\Dropbox"

    if b100:
        suffix="100"
    else:
        suffix="40"

    if bAlbum:
        suffix += "a"
    else:
        suffix += "s"
    
    if bAll:
        suffix += "_all"

    if not bCombine:
        dirname = f"{dirroot}\\Charts\\{year}"
        filename = f"{dirname}\\top{suffix}-{year}.xlsx"
    else:
        dirname = f"{dirroot}\\Charts\\{yearstart}-{yearend}"
        filename = f"{dirname}\\top{suffix}-{yearstart}-{yearend}.xlsx"

    os.makedirs(dirname, exist_ok=True)

    while os.path.isfile(filename):       
        try:
            os.remove(filename)
            removed = True
        except:
            input(f"File {filename} not accessible. Close file and press enter to retry: ")
        
    workbook = Workbook(filename)
    worksheet = workbook.add_worksheet()
    format1 = workbook.add_format({'num_format': 'dd/mm/yyyy'})
    format2 = workbook.add_format()
    format2.set_bold()

    worksheet.freeze_panes(1, 0)

    col = 0
    worksheet.write (0, 0, "START", format2);
    worksheet.set_column(0, 0, 12)
    if not bAll:
        worksheet.write (0, col:=col+1, "END", format2);
        worksheet.set_column(col, col, 12)
    if not bCombine or bAll:
        worksheet.write (0, col:=col+1, "POS", format2);
        worksheet.set_column(col, col, 8)
    worksheet.write (0, col:=col+1, "PEAK", format2);
    worksheet.set_column(col, col, 8)
    worksheet.write (0, col:=col+1, "WEEKS", format2);
    worksheet.set_column(col, col, 8)
    worksheet.write (0, col:=col+1, "TITLE", format2);
    worksheet.set_column(col, col, 50)
    worksheet.write (0, col:=col+1, "ARTIST", format2);
    worksheet.set_column(col, col, 50)
    worksheet.write (0, col:=col+1, "Heard", format2);
    worksheet.set_column(col, col, 8)

    for i, s in enumerate(parser.songs):
        col = 0
        worksheet.write(i+1, 0, s.startdate, format1)
        if not bAll:
            worksheet.write(i+1, col:=col+1, s.enddate, format1)
        if not bCombine or bAll:
            worksheet.write(i+1, col:=col+1, int(s.position))
        worksheet.write(i+1, col:=col+1, int(s.peak))
        worksheet.write(i+1, col:=col+1, int(s.weeks))
        worksheet.write(i+1, col:=col+1, s.title)
        worksheet.write(i+1, col:=col+1, s.artist)

    workbook.close()

    print (f"output file:\n{filename}") 


def sortbyweeks(item1, item2):
    if  item1.weeks == item2.weeks:
        return 1 if item1.artist > item2.artist else -1 

    return 1 if item1.weeks < item2.weeks else -1 

def sortbyartist(item1, item2):
    return 1 if item1.artist > item2.artist else -1 


def findArtist(artists, artisttofind):
    for index, a in enumerate(artists):
        if a.artist == artisttofind.artist:
            return a
    return None 

def processArtists(songs):
    artists = [] 
    songs.sort(key=functools.cmp_to_key(sortbyartist))
    for index, s in enumerate(songs):
        newArtist = artist(s.artist)
        if len(artists) == 0:
            artists.append(newArtist)
        else:
            a = findArtist (artists, newArtist)
            if a != None:
                a.weeks += 1
            else:
                artists.append(newArtist)

    print ("Artists len ",len(artists))

    artists.sort(key=functools.cmp_to_key(sortbyweeks))   
    writeartists(artists)     

def processSongs(songs, command):
    if command == "Artists":
        processArtists(songs)

def readfile():
    songs = []

    startPos = None
    endPos = None
    posPos = None
    peakPos = None
    weeksPos = None
    titlePos = None
    artistPos = None
    start = None
    end = None
    pos = None
    peak = None
    weeks = None
    title = None
    artist = None

    print("Loading ", fileToRead)
    dataframe1 = openpyxl.load_workbook(fileToRead).active
    print("Parsing ", fileToRead)
    
    min_row = dataframe1.min_row
    max_row = dataframe1.max_row
    min_col = dataframe1.min_column
    max_col = dataframe1.max_column

    for row in range(min_row - 1, max_row):
        print (row)
        if row == 0:
            for col in dataframe1.iter_cols(min_col, max_col + 1):
                value = col[row].value
                if value == "START":
                    startPos = col
                if value == "END":
                    endPos = col
                if value == "POS":
                    posPos = col
                if value == "PEAK":
                    peakPos = col
                if value == "WEEKS":
                    weeksPos = col
                if value == "TITLE":
                    titlePos = col
                if value == "ARTIST":
                    artistPos = col
        else:
            for col in dataframe1.iter_cols(1, max_col):
                q = 1
          #      value = col[row].value
          #      if col == startPos:
           #         start = value
            #    if col == endPos:
             #       end = value
              #  if col == posPos:
            #        pos = value
             #   if col == peakPos:
              #      peak = value
             #   if col == weeksPos:
             #       weeks = value
             #   if col == titlePos:
             #       title = value
             #   if col == artistPos:
             #       artist = value            
     
       # if ( row % 10 == 0):
        #    print(row, artist)

        #if title != None:
        #    songs.append(song(title, artist, pos, peak, weeks, start, end))
 
    print(f"Parsed {len(songs)} lines")

    if cmd != None:
        processSongs(songs, cmd)

parser = MyHTMLParser()
fileToRead = ""
bCombine = False
bAlbum = False
b100 = False
week1 = 0
week2 = 52
bAll = False
year = 0
yearstart = 0
yearend = 0
cmd = None

if len(sys.argv) > 1:
    if sys.argv[1].isnumeric() and int(sys.argv[1]) > 1950:
        yearstart = int(sys.argv[1])
        yearend = yearstart
    
    if (len(sys.argv) > 2 and sys.argv[2].isnumeric() and int(sys.argv[2]) > 1950):
        yearend = int(sys.argv[2])

    for arg in range (1, len(sys.argv)):
        if sys.argv[arg] == "r":
            fileToRead = sys.argv[arg+1] 
        if sys.argv[arg] == "c":
            bCombine = True
        if sys.argv[arg] == "a":
            bAlbum = True
        if sys.argv[arg] == "w1":
            week1 = int(sys.argv[arg+1]) - 1
        if sys.argv[arg] == "w2":
            week2 = int(sys.argv[arg+1])
        if sys.argv[arg] == "100":
            b100 = True
        if sys.argv[arg] == "all":
            bAll = True
        if sys.argv[arg] == "cmd":
            cmd = sys.argv[arg+1]

    if fileToRead == "":
        for year in range ( yearstart, yearend+1):
            dtstart = datetime.datetime.strptime(f"{year}0107", '%Y%m%d')
            for week in  range (week1, week2):
                dtnow = dtstart + timedelta(weeks=week)
                print(dtnow)
                if bAlbum:
                    url = f"https://www.officialcharts.com/charts/singles-chart/{dtnow.strftime('%Y%m%d')}/7502/"
                else:
                    if b100:
                        url = f"https://www.officialcharts.com/charts/singles-chart/{dtnow.strftime('%Y%m%d')}/7501/"
                    else:
                        url = f"https://www.officialcharts.com/charts/singles-chart/{dtnow.strftime('%Y%m%d')}/750140/"

                fp = urllib.request.urlopen(url)
                mystr = fp.read().decode("utf8")
                parser.chartdate = dtnow
                parser.feed(mystr)
            
            if not bCombine:
                writefile()
                parser.clear()
        if bCombine:
            writefile()
    else:
        readfile()
else:
    print ( "Usage: charts <year>")

    

